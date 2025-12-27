from django.core.management.base import BaseCommand
from datetime import datetime, timedelta
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from management.database import data_mysql
from management.utils import _format_idr_number, send_telegram_document_aiogram, send_telegram_message_aiogram


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('--tanggal', type=str, default=None)

    def handle(self, *args, **kwargs):
        chat_id = (os.getenv('TELEGRAM_EARNING_REPORT_CHAT_ID') or os.getenv('TELEGRAM_DEFAULT_CHAT_ID') or '').strip()
        if not chat_id:
            return

        tanggal = (kwargs.get('tanggal') or '').strip()
        if tanggal:
            today = datetime.strptime(tanggal, '%Y-%m-%d').date()
        else:
            today = datetime.now().date()

        start_date = today
        end_date = today
        past_start_date = today - timedelta(days=1)
        past_end_date = past_start_date

        res = data_mysql().get_all_rekapitulasi_adx_monitoring_account_by_params(
            start_date,
            end_date,
            past_start_date,
            past_end_date,
            [],
            [],
        )
        rows = ((res or {}).get('hasil') or {}).get('data') or []

        def rp(v):
            return _format_idr_number(v).replace('Rp ', 'Rp. ')

        total_now = 0
        total_last = 0
        for r in rows:
            try:
                total_now += float(r.get('pendapatan_now') or 0)
            except Exception:
                pass
            try:
                total_last += float(r.get('pendapatan_last') or 0)
            except Exception:
                pass

        generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title = f"Rekapitulasi Earning Harian ({today.strftime('%Y-%m-%d')})"
        summary_msg = (
            f"{title}\n"
            f"Total Hari Ini: {rp(total_now)}\n"
            f"Total Kemarin: {rp(total_last)}\n"
            f"Waktu: {generated_at}"
        )

        send_telegram_message_aiogram(chat_id, summary_msg)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Earning Harian'

        ws['A1'] = title
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = 'Waktu generate'
        ws['B2'] = generated_at

        headers = ['Account', 'Pendapatan Hari Ini', 'Pendapatan Kemarin', 'Selisih', 'Persen (%)']
        header_row = 4
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        rrow = header_row + 1
        for r in rows:
            acc_name = str(r.get('account_name') or r.get('account_id') or '-')
            now_v = r.get('pendapatan_now')
            last_v = r.get('pendapatan_last')
            diff_v = r.get('pendapatan_selisih')
            pct_v = r.get('pendapatan_persen')

            ws.cell(row=rrow, column=1, value=acc_name)
            ws.cell(row=rrow, column=2, value=rp(now_v))
            ws.cell(row=rrow, column=3, value=rp(last_v))
            ws.cell(row=rrow, column=4, value=rp(diff_v))
            ws.cell(row=rrow, column=5, value=(float(pct_v) if pct_v not in [None, ''] else None))
            rrow += 1

        ws.cell(row=rrow, column=1, value='Total').font = Font(bold=True)
        ws.cell(row=rrow, column=2, value=rp(total_now)).font = Font(bold=True)
        ws.cell(row=rrow, column=3, value=rp(total_last)).font = Font(bold=True)
        ws.cell(row=rrow, column=4, value=rp(total_now - total_last)).font = Font(bold=True)

        widths = [40, 22, 22, 22, 12]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        buf = BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        filename = f"rekapitulasi_earning_{today.strftime('%Y-%m-%d')}.xlsx"
        send_telegram_document_aiogram(chat_id, xlsx_bytes, filename, caption=title)