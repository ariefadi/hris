from django.core.management.base import BaseCommand
from datetime import datetime, timedelta
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from management.database import data_mysql
from management.utils import _format_idr_number, send_telegram_document_aiogram, send_telegram_message_aiogram


def _extract_base_subdomain(full_string):
    s = str(full_string or '').strip()
    if not s:
        return ''
    parts = s.split('.')
    if len(parts) >= 2:
        return '.'.join(parts[:2])
    return s


def _safe_float(v, default=0.0):
    try:
        return float(v or 0)
    except Exception:
        return float(default)


def _safe_str(v, default=''):
    try:
        s = str(v or '').strip()
        return s if s else default
    except Exception:
        return default


def _pick_account_name(accounts_spend):
    if not isinstance(accounts_spend, dict) or not accounts_spend:
        return ''
    best = None
    best_spend = -1.0
    for k, v in accounts_spend.items():
        name = _safe_str(k, '')
        spend = _safe_float(v, 0.0)
        if not name:
            continue
        if spend > best_spend:
            best_spend = spend
            best = name
    return best or ''


def _combine_roi_domain(adx_rows, fb_rows):
    fb_map = {}
    for fb_item in (fb_rows or []):
        domain = _safe_str(fb_item.get('domain'), '')
        base_domain = _extract_base_subdomain(domain)
        cc = _safe_str(fb_item.get('country_code'), '').upper()
        if not base_domain or not cc:
            continue
        key = f"{base_domain}_{cc}"
        fb_map[key] = fb_item

    grouped_all = {}
    grouped_filtered = {}

    for adx_item in (adx_rows or []):
        subdomain = _safe_str(adx_item.get('site_name'), '')
        if not subdomain:
            continue
        base_subdomain = _extract_base_subdomain(subdomain)
        cc = _safe_str(adx_item.get('country_code'), '').upper()
        if not base_subdomain or not cc:
            continue

        fb_data = fb_map.get(f"{base_subdomain}_{cc}") or {}
        account_ads = _safe_str(fb_data.get('account_name'), '')
        spend = _safe_float(fb_data.get('spend'), 0.0)
        revenue = _safe_float(adx_item.get('revenue'), 0.0)

        cur = grouped_all.get(subdomain)
        if not cur:
            grouped_all[subdomain] = {
                'site_name': subdomain,
                'spend': 0.0,
                'revenue': 0.0,
                'accounts_spend': {},
            }
            cur = grouped_all[subdomain]
        cur['spend'] += spend
        cur['revenue'] += revenue
        if account_ads:
            cur['accounts_spend'][account_ads] = _safe_float(cur['accounts_spend'].get(account_ads), 0.0) + spend

        if spend > 0:
            curf = grouped_filtered.get(subdomain)
            if not curf:
                grouped_filtered[subdomain] = {
                    'site_name': subdomain,
                    'spend': 0.0,
                    'revenue': 0.0,
                    'accounts_spend': {},
                }
                curf = grouped_filtered[subdomain]
            curf['spend'] += spend
            curf['revenue'] += revenue
            if account_ads:
                curf['accounts_spend'][account_ads] = _safe_float(curf['accounts_spend'].get(account_ads), 0.0) + spend

    items_all = []
    for v in grouped_all.values():
        spend = _safe_float(v.get('spend'), 0.0)
        revenue = _safe_float(v.get('revenue'), 0.0)
        roi = ((revenue - spend) / spend * 100) if spend > 0 else 0.0
        items_all.append(
            {
                'site_name': v.get('site_name') or '-',
                'account_ads': _pick_account_name(v.get('accounts_spend')) or '',
                'spend': round(spend, 2),
                'revenue': round(revenue, 2),
                'roi': round(roi, 2),
            }
        )

    items_filtered = []
    total_spend_filtered = 0.0
    total_revenue_filtered = 0.0
    for v in grouped_filtered.values():
        spend = _safe_float(v.get('spend'), 0.0)
        revenue = _safe_float(v.get('revenue'), 0.0)
        roi = ((revenue - spend) / spend * 100) if spend > 0 else 0.0
        items_filtered.append(
            {
                'site_name': v.get('site_name') or '-',
                'account_ads': _pick_account_name(v.get('accounts_spend')) or '',
                'spend': round(spend, 2),
                'revenue': round(revenue, 2),
                'roi': round(roi, 2),
            }
        )
        total_spend_filtered += spend
        total_revenue_filtered += revenue

    items_all.sort(key=lambda x: float(x.get('roi') or 0), reverse=True)
    items_filtered.sort(key=lambda x: float(x.get('roi') or 0), reverse=True)

    roi_nett_filtered = (
        ((total_revenue_filtered - total_spend_filtered) / total_spend_filtered * 100)
        if total_spend_filtered > 0
        else 0.0
    )

    return {
        'items_all': items_all,
        'items_filtered': items_filtered,
        'summary_filtered': {
            'total_spend': round(total_spend_filtered, 2),
            'total_revenue': round(total_revenue_filtered, 2),
            'roi_nett': round(roi_nett_filtered, 2),
        },
    }


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('--tanggal', type=str, default=None)

    def handle(self, *args, **kwargs):
        chat_id = (
            (os.getenv('TELEGRAM_ROI_DOMAIN_REPORT_CHAT_ID') or '').strip()
            or (os.getenv('TELEGRAM_DEFAULT_CHAT_ID') or '').strip()
        )
        if not chat_id:
            return

        tanggal = (kwargs.get('tanggal') or '').strip()
        if tanggal:
            today = datetime.strptime(tanggal, '%Y-%m-%d').date()
        else:
            today = datetime.now().date()

        yesterday = today - timedelta(days=1)
        today_str = today.strftime('%Y-%m-%d')
        yest_str = yesterday.strftime('%Y-%m-%d')

        def rp(v):
            return _format_idr_number(v).replace('Rp ', 'Rp. ')

        db = data_mysql()

        adx_today_res = db.get_all_adx_monitoring_account_by_params(today_str, today_str, [], [])
        adx_today_rows = ((adx_today_res or {}).get('hasil') or {}).get('data') or []

        unique_name_site = []
        try:
            sset = set()
            for r in adx_today_rows:
                site = _safe_str(r.get('site_name'), '')
                base = _extract_base_subdomain(site)
                if base and base != 'Unknown':
                    sset.add(base)
            unique_name_site = list(sset)
        except Exception:
            unique_name_site = []

        fb_today_rows = []
        if unique_name_site:
            fb_today_res = db.get_all_ads_roi_monitoring_campaign_by_params(today_str, today_str, unique_name_site)
            fb_today_rows = ((fb_today_res or {}).get('hasil') or {}).get('data') or []

        combined_today = _combine_roi_domain(adx_today_rows, fb_today_rows)
        roi_today = _safe_float(((combined_today.get('summary_filtered') or {}).get('roi_nett')), 0.0)

        adx_yest_res = db.get_all_adx_monitoring_account_by_params(yest_str, yest_str, [], [])
        adx_yest_rows = ((adx_yest_res or {}).get('hasil') or {}).get('data') or []

        fb_yest_rows = []
        if unique_name_site:
            fb_yest_res = db.get_all_ads_roi_monitoring_campaign_by_params(yest_str, yest_str, unique_name_site)
            fb_yest_rows = ((fb_yest_res or {}).get('hasil') or {}).get('data') or []

        combined_yest = _combine_roi_domain(adx_yest_rows, fb_yest_rows)
        roi_yest = _safe_float(((combined_yest.get('summary_filtered') or {}).get('roi_nett')), 0.0)

        delta = roi_today - roi_yest
        delta_dir = 'Naik' if delta >= 0 else 'Turun'
        delta_abs = abs(delta)

        top_candidates = list(combined_today.get('items_filtered') or [])
        top_by_roi = sorted(top_candidates, key=lambda x: float(x.get('roi') or 0), reverse=True)[:10]
        top_by_roi_then_rev = sorted(top_by_roi, key=lambda x: float(x.get('revenue') or 0), reverse=True)

        top_names = []
        for x in top_by_roi_then_rev:
            dom = _safe_str(x.get('site_name'), '-')
            acc = _safe_str(x.get('account_ads'), '')
            if acc:
                top_names.append(f"{dom} ( {acc} )")
            else:
                top_names.append(dom)

        generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title = f"ROI Per Domain Harian ({today_str})"
        msg = (
            f"{title}\n"
            f"ROI Hari Ini Sebesar : {roi_today:.2f}%\n"
            f"ROI Kemarin Sebesar: Rp. {roi_yest:.2f} %\n"
            f"ROI ({delta_dir}) Sebesar : {delta_abs:.2f}%\n"
            f"10 Domain Dengan ROI Tertinggi Berdasarkan Urutan Pendapatan Tertinggi Adalah : {', '.join(top_names) if top_names else '-'}\n\n"
            f"Waktu: {generated_at}"
        )
        send_telegram_message_aiogram(chat_id, msg)

        rows_excel = sorted(
            (combined_today.get('items_filtered') or []),
            key=lambda x: float(x.get('revenue') or 0),
            reverse=True,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = 'ROI Per Domain'

        ws['A1'] = title
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = 'Waktu generate'
        ws['B2'] = generated_at

        headers = ['Subdomain', 'Account Ads', 'Spend (Rp)', 'Pendapatan (Rp)', 'ROI (%)']
        header_row = 4
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        rrow = header_row + 1
        for r in rows_excel:
            ws.cell(row=rrow, column=1, value=_safe_str(r.get('site_name'), '-'))
            ws.cell(row=rrow, column=2, value=_safe_str(r.get('account_ads'), '-'))
            ws.cell(row=rrow, column=3, value=rp(r.get('spend')))
            ws.cell(row=rrow, column=4, value=rp(r.get('revenue')))
            ws.cell(row=rrow, column=5, value=float(r.get('roi') or 0.0))
            rrow += 1

        widths = [38, 26, 18, 20, 12]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        buf = BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        filename = f"roi_per_domain_{today_str}.xlsx"
        send_telegram_document_aiogram(chat_id, xlsx_bytes, filename, caption=title)