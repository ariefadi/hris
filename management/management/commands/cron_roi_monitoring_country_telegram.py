from django.core.management.base import BaseCommand
from datetime import datetime, timedelta
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from management.database import data_mysql
from management.utils import _format_idr_number, send_telegram_document_aiogram, send_telegram_message_aiogram


def _extract_base_subdomain(full_string):
    s = str(full_string or '').strip()
    if not s:
        return ''
    parts = s.split('.')
    if len(parts) >= 2:
        return '.'.join(parts[:2])
    return s


def _flag_emoji(country_code):
    cc = (str(country_code or '').strip().upper())
    if len(cc) != 2 or (not cc.isalpha()):
        return ''
    base = 0x1F1E6
    return chr(base + ord(cc[0]) - 65) + chr(base + ord(cc[1]) - 65)


def _safe_float(v, default=0.0):
    try:
        return float(v or 0)
    except Exception:
        return float(default)


def _combine_roi_country(adx_rows, fb_rows):
    adx_map = {}
    fb_map = {}
    country_name_by_code = {}

    for r in (adx_rows or []):
        date_key = str(r.get('date') or '').strip()
        site = str(r.get('site_name') or '').strip()
        base_site = _extract_base_subdomain(site)
        cc = (str(r.get('country_code') or '').strip().upper())
        cn = str(r.get('country_name') or '').strip()
        revenue = _safe_float(r.get('revenue'), 0.0)
        if not date_key or not base_site or not cc:
            continue
        if cn:
            country_name_by_code[cc] = cn
        key = f"{date_key}_{base_site}_{cc}"
        adx_map[key] = (adx_map.get(key, 0.0) + revenue)

    for r in (fb_rows or []):
        date_key = str(r.get('date') or '').strip()
        domain = str(r.get('domain') or '').strip()
        base_dom = _extract_base_subdomain(domain)
        cc = (str(r.get('country_code') or '').strip().upper())
        cn = str(r.get('country_name') or '').strip()
        spend = _safe_float(r.get('spend'), 0.0)
        if not date_key or not base_dom or not cc:
            continue
        if cn:
            country_name_by_code[cc] = cn
        key = f"{date_key}_{base_dom}_{cc}"
        fb_map[key] = (fb_map.get(key, 0.0) + spend)

    agg_all = {}
    agg_filtered = {}
    for key in set(list(adx_map.keys()) + list(fb_map.keys())):
        try:
            cc = key.split('_')[-1]
        except Exception:
            continue
        if not cc:
            continue
        revenue = _safe_float(adx_map.get(key), 0.0)
        spend = _safe_float(fb_map.get(key), 0.0)
        name = country_name_by_code.get(cc, '') or cc

        cur_all = agg_all.get(cc)
        if not cur_all:
            agg_all[cc] = {'country': name, 'country_code': cc, 'spend': 0.0, 'revenue': 0.0}
            cur_all = agg_all[cc]
        cur_all['spend'] += spend
        cur_all['revenue'] += revenue

        if spend > 0:
            cur_f = agg_filtered.get(cc)
            if not cur_f:
                agg_filtered[cc] = {'country': name, 'country_code': cc, 'spend': 0.0, 'revenue': 0.0}
                cur_f = agg_filtered[cc]
            cur_f['spend'] += spend
            cur_f['revenue'] += revenue

    items_all = []
    for cc, v in agg_all.items():
        spend = _safe_float(v.get('spend'), 0.0)
        revenue = _safe_float(v.get('revenue'), 0.0)
        roi = ((revenue - spend) / spend * 100) if spend > 0 else 0.0
        items_all.append(
            {
                'country': v.get('country') or cc,
                'country_code': cc,
                'spend': round(spend, 2),
                'revenue': round(revenue, 2),
                'roi': round(roi, 2),
            }
        )

    items_filtered = []
    for cc, v in agg_filtered.items():
        spend = _safe_float(v.get('spend'), 0.0)
        revenue = _safe_float(v.get('revenue'), 0.0)
        roi = ((revenue - spend) / spend * 100) if spend > 0 else 0.0
        items_filtered.append(
            {
                'country': v.get('country') or cc,
                'country_code': cc,
                'spend': round(spend, 2),
                'revenue': round(revenue, 2),
                'roi': round(roi, 2),
            }
        )

    items_all.sort(key=lambda x: float(x.get('roi') or 0), reverse=True)
    items_filtered.sort(key=lambda x: float(x.get('roi') or 0), reverse=True)

    total_spend_filtered = sum(_safe_float(x.get('spend'), 0.0) for x in items_filtered)
    total_revenue_filtered = sum(_safe_float(x.get('revenue'), 0.0) for x in items_filtered)
    roi_nett_filtered = ((total_revenue_filtered - total_spend_filtered) / total_spend_filtered * 100) if total_spend_filtered > 0 else 0.0

    return {
        'items_all': items_all,
        'items_filtered': items_filtered,
        'summary_filtered': {
            'total_spend': round(total_spend_filtered, 2),
            'total_revenue': round(total_revenue_filtered, 2),
            'roi_nett': round(roi_nett_filtered, 2),
        },
    }


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('--tanggal', type=str, default=None)

    def handle(self, *args, **kwargs):
        chat_id = (
            (os.getenv('TELEGRAM_ROI_COUNTRY_REPORT_CHAT_ID') or '').strip()
            or (os.getenv('TELEGRAM_DEFAULT_CHAT_ID') or '').strip()
        )
        if not chat_id:
            return

        tanggal = (kwargs.get('tanggal') or '').strip()
        if tanggal:
            today = datetime.strptime(tanggal, '%Y-%m-%d').date()
        else:
            today = datetime.now().date()

        yesterday = today - timedelta(days=1)
        today_str = today.strftime('%Y-%m-%d')
        yest_str = yesterday.strftime('%Y-%m-%d')

        def rp(v):
            return _format_idr_number(v).replace('Rp ', 'Rp. ')

        db = data_mysql()

        adx_today = db.get_all_adx_country_detail_by_params(today_str, today_str, [], [], [])
        adx_today_rows = (adx_today or {}).get('data') if isinstance(adx_today, dict) else []
        sites = []
        try:
            unique_sites = set()
            for r in (adx_today_rows or []):
                s = _extract_base_subdomain(r.get('site_name'))
                if s and s != 'Unknown':
                    unique_sites.add(s)
            sites = list(unique_sites)
        except Exception:
            sites = []

        fb_today_rows = []
        if sites:
            fb_today_res = db.get_all_ads_country_detail_by_params(today_str, today_str, sites, [])
            fb_today_rows = ((fb_today_res or {}).get('hasil') or {}).get('data') or []

        combined_today = _combine_roi_country(adx_today_rows, fb_today_rows)
        roi_today = _safe_float(((combined_today.get('summary_filtered') or {}).get('roi_nett')), 0.0)

        adx_yest = db.get_all_adx_country_detail_by_params(yest_str, yest_str, [], [], [])
        adx_yest_rows = (adx_yest or {}).get('data') if isinstance(adx_yest, dict) else []
        fb_yest_rows = []
        if sites:
            fb_yest_res = db.get_all_ads_country_detail_by_params(yest_str, yest_str, sites, [])
            fb_yest_rows = ((fb_yest_res or {}).get('hasil') or {}).get('data') or []

        combined_yest = _combine_roi_country(adx_yest_rows, fb_yest_rows)
        roi_yest = _safe_float(((combined_yest.get('summary_filtered') or {}).get('roi_nett')), 0.0)

        delta = roi_today - roi_yest
        delta_dir = 'Naik' if delta >= 0 else 'Turun'
        delta_abs = abs(delta)

        rows_excel = sorted(
            (combined_today.get('items_filtered') or []),
            key=lambda x: float(x.get('revenue') or 0),
            reverse=True,
        )

        top_rows = rows_excel[:10]
        top_names = []
        for x in top_rows:
            cc = x.get('country_code')
            cn = x.get('country') or cc or '-'
            fl = _flag_emoji(cc)
            if fl:
                top_names.append(f"{fl} {cn}")
            else:
                top_names.append(str(cn))

        generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title = f"ROI Per Negara Harian ({today_str})"
        msg = (
            f"{title}\n"
            f"ROI Hari Ini Sebesar : {roi_today:.2f}%\n"
            f"ROI Kemarin Sebesar: Rp. {roi_yest:.2f} %\n"
            f"ROI ({delta_dir}) Sebesar : {delta_abs:.2f}%\n"
            f"10 Negara Dengan ROI Tertinggi Berdasarkan Urutan Pendapatan Tertinggi Adalah : {', '.join(top_names) if top_names else '-'}\n"
            f"Waktu: {generated_at}"
        )
        send_telegram_message_aiogram(chat_id, msg)

        wb = Workbook()
        ws = wb.active
        ws.title = 'ROI Per Negara'

        ws['A1'] = title
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = 'Waktu generate'
        ws['B2'] = generated_at

        headers = ['Negara', 'Kode Negara', 'Spend (Rp)', 'Pendapatan (Rp)', 'ROI (%)']
        header_row = 4
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        rrow = header_row + 1
        for r in rows_excel:
            ws.cell(row=rrow, column=1, value=str(r.get('country') or '-'))
            ws.cell(row=rrow, column=2, value=str(r.get('country_code') or '-'))
            ws.cell(row=rrow, column=3, value=rp(r.get('spend')))
            ws.cell(row=rrow, column=4, value=rp(r.get('revenue')))
            ws.cell(row=rrow, column=5, value=float(r.get('roi') or 0.0))
            rrow += 1

        widths = [30, 14, 18, 20, 12]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        buf = BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        filename = f"roi_per_country_{today_str}.xlsx"
        send_telegram_document_aiogram(chat_id, xlsx_bytes, filename, caption=title)